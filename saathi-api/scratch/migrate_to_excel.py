import sqlite3
import os
import json
from pathlib import Path
import openpyxl

def migrate():
    sqlite_file = "saathi.db"
    excel_file = "saathi_data.xlsx"
    
    if not os.path.exists(sqlite_file):
        print("SQLite file not found.")
        return
        
    conn = sqlite3.connect(sqlite_file)
    conn.row_factory = sqlite3.Row
    
    wb = openpyxl.load_workbook(excel_file)
    
    # Migrate Profiles
    cur = conn.cursor()
    cur.execute("SELECT * FROM profiles")
    profiles = [dict(r) for r in cur.fetchall()]
    ws = wb["Profiles"]
    headers = [cell.value for cell in ws[1]]
    for p in profiles:
        row = [p.get(h) for h in headers]
        ws.append(row)
        
    # Migrate Sessions
    cur.execute("SELECT * FROM chat_sessions_v2")
    sessions = [dict(r) for r in cur.fetchall()]
    ws = wb["Sessions"]
    headers = [cell.value for cell in ws[1]]
    for s in sessions:
        row = [s.get(h) for h in headers]
        ws.append(row)
        
    # Migrate Messages
    cur.execute("SELECT * FROM chat_messages_v2")
    messages = [dict(r) for r in cur.fetchall()]
    ws = wb["Messages"]
    headers = [cell.value for cell in ws[1]]
    for m in messages:
        row = [m.get(h) for h in headers]
        ws.append(row)
        
    # Migrate Integrations
    cur.execute("SELECT * FROM google_integrations")
    integrations = [dict(r) for r in cur.fetchall()]
    ws = wb["Integrations"]
    headers = [cell.value for cell in ws[1]]
    for i in integrations:
        row = [i.get(h) for h in headers]
        ws.append(row)
        
    wb.save(excel_file)
    print("Migration complete!")

if __name__ == "__main__":
    migrate()
