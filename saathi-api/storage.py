import datetime as dt
import json
import os
import sqlite3
import uuid
import random
from pathlib import Path
from typing import Any, Optional

import httpx
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

LOCAL_DB_FILE = Path(__file__).with_name("saathi.db")
SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip().rstrip("/")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "").strip()
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()


def _now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def supabase_enabled() -> bool:
    return bool(SUPABASE_URL and SUPABASE_ANON_KEY and SUPABASE_SERVICE_ROLE_KEY)


class StorageError(RuntimeError):
    pass


class LocalStore:
    def __init__(self, db_file: Path):
        self.db_file = db_file
        self.init()

    def _connect(self):
        conn = sqlite3.connect(self.db_file)
        conn.row_factory = sqlite3.Row
        return conn

    def init(self):
        conn = self._connect()
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS profiles (
                id TEXT PRIMARY KEY,
                email TEXT,
                full_name TEXT,
                language TEXT DEFAULT 'English',
                voice_gender TEXT DEFAULT 'female',
                onboarding_completed INTEGER DEFAULT 0,
                created_at TEXT,
                updated_at TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS chat_sessions_v2 (
                id TEXT PRIMARY KEY,
                user_id TEXT NOT NULL,
                title TEXT NOT NULL,
                created_at TEXT,
                updated_at TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS chat_messages_v2 (
                id TEXT PRIMARY KEY,
                session_id TEXT NOT NULL,
                user_id TEXT NOT NULL,
                role TEXT NOT NULL,
                text TEXT NOT NULL,
                created_at TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS google_integrations (
                user_id TEXT PRIMARY KEY,
                google_email TEXT,
                access_token TEXT,
                refresh_token TEXT,
                token_expiry TEXT,
                scopes_json TEXT,
                connected_at TEXT,
                updated_at TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS nudges (
                id TEXT PRIMARY KEY,
                user_id TEXT NOT NULL,
                text TEXT NOT NULL,
                detail TEXT,
                type TEXT DEFAULT 'deadline',
                status TEXT DEFAULT 'active',
                created_at TEXT
            )
            """
        )
    def _doctor(self):
        """Forces correct schema for v29+ tables."""
        conn = self._connect()
        cur = conn.cursor()
        
        # Check if nudges is legacy (INTEGER id)
        cur.execute("PRAGMA table_info(nudges)")
        info = {row['name']: row for row in cur.fetchall()}
        
        if 'id' in info and info['id']['type'] == 'INTEGER':
            print("🚀 Migrating legacy nudges table...")
            cur.execute("DROP TABLE nudges")
            
        cur.execute("""
            CREATE TABLE IF NOT EXISTS nudges (
                id TEXT PRIMARY KEY,
                user_id TEXT NOT NULL,
                text TEXT NOT NULL,
                detail TEXT,
                type TEXT DEFAULT 'deadline',
                status TEXT DEFAULT 'active',
                created_at TEXT
            )
        """)
        
        cur.execute("CREATE TABLE IF NOT EXISTS memory_snippets (id TEXT PRIMARY KEY, user_id TEXT NOT NULL, text TEXT NOT NULL, created_at TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS deadlines (id TEXT PRIMARY KEY, user_id TEXT NOT NULL, title TEXT NOT NULL, due_date TEXT, source_email_id TEXT, status TEXT DEFAULT 'pending', created_at TEXT)")
        
        # Ensure correct columns for memory_snippets just in case
        try: cur.execute("ALTER TABLE memory_snippets ADD COLUMN created_at TEXT")
        except: pass
        
        try: cur.execute("ALTER TABLE deadlines ADD COLUMN source_email_id TEXT")
        except: pass
        try: cur.execute("ALTER TABLE chat_sessions_v2 ADD COLUMN title TEXT")
        except: pass

        conn.commit()
        conn.close()

    def list_memory(self, user_id: str) -> list[dict[str, Any]]:
        self._doctor()
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT * FROM memory_snippets WHERE user_id = ? ORDER BY created_at DESC", (user_id,))
        rows = [dict(row) for row in cur.fetchall()]
        conn.close()
        return rows

    def add_memory(self, user_id: str, text: str) -> dict[str, Any]:
        self._doctor()
        m_id = str(uuid.uuid4())
        now = _now_iso()
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("INSERT INTO memory_snippets (id, user_id, text, created_at) VALUES (?, ?, ?, ?)", (m_id, user_id, text, now))
        conn.commit()
        conn.close()
        return {"id": m_id, "text": text, "created_at": now}

    def delete_memory(self, user_id: str, memory_id: str):
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("DELETE FROM memory_snippets WHERE user_id = ? AND id = ?", (user_id, memory_id))
        conn.commit()
        conn.close()

    def list_deadlines(self, user_id: str) -> list[dict[str, Any]]:
        self._doctor()
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT * FROM deadlines WHERE user_id = ? ORDER BY due_date ASC", (user_id,))
        rows = [dict(row) for row in cur.fetchall()]
        conn.close()
        return rows

    def add_deadline(self, user_id: str, title: str, due_date: str, source_email_id: str = "") -> dict[str, Any]:
        self._doctor()
        conn = self._connect()
        cur = conn.cursor()
        # Existence Check
        cur.execute("SELECT id FROM deadlines WHERE user_id = ? AND title = ? AND due_date = ?", (user_id, title, due_date))
        if cur.fetchone():
            conn.close()
            return {"status": "exists"}

        d_id = str(uuid.uuid4())
        now = _now_iso()
        cur.execute(
            "INSERT INTO deadlines (id, user_id, title, due_date, source_email_id, status, created_at) VALUES (?, ?, ?, ?, ?, 'pending', ?)",
            (d_id, user_id, title, due_date, source_email_id, now)
        )
        conn.commit()
        conn.close()
        return {"id": d_id, "title": title, "due_date": due_date, "status": "pending"}

    def update_deadline_status(self, user_id: str, deadline_id: str, status: str):
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("UPDATE deadlines SET status = ? WHERE user_id = ? AND id = ?", (status, user_id, deadline_id))
        conn.commit()
        conn.close()

    def delete_deadline(self, user_id: str, deadline_id: str):
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("DELETE FROM deadlines WHERE user_id = ? AND id = ?", (user_id, deadline_id))
        conn.commit()
        conn.close()

    def ensure_profile(self, user_id: str, email: str = "", full_name: str = "") -> dict[str, Any]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT * FROM profiles WHERE id = ?", (user_id,))
        row = cur.fetchone()
        if not row:
            now = _now_iso()
            cur.execute(
                """
                INSERT INTO profiles (id, email, full_name, language, voice_gender, onboarding_completed, created_at, updated_at)
                VALUES (?, ?, ?, 'English', 'female', 0, ?, ?)
                """,
                (user_id, email, full_name or "Friend", now, now),
            )
            conn.commit()
            cur.execute("SELECT * FROM profiles WHERE id = ?", (user_id,))
            row = cur.fetchone()
        profile = dict(row)
        profile["onboarding_completed"] = bool(profile.get("onboarding_completed"))
        conn.close()
        return profile

    def update_profile(self, user_id: str, data: dict[str, Any]) -> dict[str, Any]:
        allowed = {"email", "full_name", "language", "voice_gender", "onboarding_completed"}
        payload = {key: value for key, value in data.items() if key in allowed}
        if not payload:
            return self.ensure_profile(user_id)

        assignments = ", ".join(f"{key} = ?" for key in payload)
        values = list(payload.values()) + [_now_iso(), user_id]
        conn = self._connect()
        cur = conn.cursor()
        cur.execute(
            f"UPDATE profiles SET {assignments}, updated_at = ? WHERE id = ?",
            values,
        )
        conn.commit()
        conn.close()
        return self.ensure_profile(user_id)

    def list_sessions(self, user_id: str) -> list[dict[str, Any]]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT * FROM chat_sessions_v2
            WHERE user_id = ?
            ORDER BY updated_at DESC
            """,
            (user_id,),
        )
        rows = [dict(row) for row in cur.fetchall()]
        conn.close()
        return rows

    def create_session(self, user_id: str, title: str = "New chat") -> dict[str, Any]:
        session_id = str(uuid.uuid4())
        now = _now_iso()
        conn = self._connect()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO chat_sessions_v2 (id, user_id, title, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (session_id, user_id, title, now, now),
        )
        conn.commit()
        conn.close()
        return {"id": session_id, "user_id": user_id, "title": title, "created_at": now, "updated_at": now}

    def get_messages(self, user_id: str, session_id: str) -> list[dict[str, Any]]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT role, text, created_at, id
            FROM chat_messages_v2
            WHERE user_id = ? AND session_id = ?
            ORDER BY created_at ASC, id ASC
            """,
            (user_id, session_id),
        )
        rows = [dict(row) for row in cur.fetchall()]
        conn.close()
        return rows

    def add_message(self, user_id: str, session_id: str, role: str, text: str) -> dict[str, Any]:
        message_id = str(uuid.uuid4())
        now = _now_iso()
        conn = self._connect()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO chat_messages_v2 (id, session_id, user_id, role, text, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (message_id, session_id, user_id, role, text, now),
        )
        cur.execute(
            "UPDATE chat_sessions_v2 SET updated_at = ? WHERE id = ? AND user_id = ?",
            (now, session_id, user_id),
        )
        if role == "user":
            cur.execute(
                """
                SELECT COUNT(*)
                FROM chat_messages_v2
                WHERE user_id = ? AND session_id = ? AND role = 'user'
                """,
                (user_id, session_id),
            )
            if cur.fetchone()[0] == 1:
                title = (text[:42].strip() + "...") if len(text.strip()) > 45 else text.strip() or "New chat"
                cur.execute(
                    "UPDATE chat_sessions_v2 SET title = ? WHERE id = ? AND user_id = ?",
                    (title, session_id, user_id),
                )
        conn.commit()
        conn.close()
        return {"id": message_id, "session_id": session_id, "user_id": user_id, "role": role, "text": text, "created_at": now}

    def delete_session(self, user_id: str, session_id: str):
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("DELETE FROM chat_messages_v2 WHERE user_id = ? AND session_id = ?", (user_id, session_id))
        cur.execute("DELETE FROM chat_sessions_v2 WHERE user_id = ? AND id = ?", (user_id, session_id))
        conn.commit()
        conn.close()

    def get_google_integration(self, user_id: str) -> Optional[dict[str, Any]]:
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT * FROM google_integrations WHERE user_id = ?", (user_id,))
        row = cur.fetchone()
        conn.close()
        if not row:
            return None
        data = dict(row)
        data["scopes"] = json.loads(data.get("scopes_json") or "[]")
        return data

    def upsert_google_integration(self, user_id: str, data: dict[str, Any]) -> dict[str, Any]:
        now = _now_iso()
        existing = self.get_google_integration(user_id)
        scopes = json.dumps(data.get("scopes") or [])
        conn = self._connect()
        cur = conn.cursor()
        if existing:
            cur.execute(
                """
                UPDATE google_integrations
                SET google_email = ?, access_token = ?, refresh_token = ?, token_expiry = ?, scopes_json = ?, updated_at = ?
                WHERE user_id = ?
                """,
                (
                    data.get("google_email"),
                    data.get("access_token"),
                    data.get("refresh_token"),
                    data.get("token_expiry"),
                    scopes,
                    now,
                    user_id,
                ),
            )
        else:
            cur.execute(
                """
                INSERT INTO google_integrations
                (user_id, google_email, access_token, refresh_token, token_expiry, scopes_json, connected_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    user_id,
                    data.get("google_email"),
                    data.get("access_token"),
                    data.get("refresh_token"),
                    data.get("token_expiry"),
                    scopes,
                    now,
                    now,
                ),
            )
        conn.commit()
        conn.close()
        return self.get_google_integration(user_id) or {}

    def delete_google_integration(self, user_id: str):
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("DELETE FROM google_integrations WHERE user_id = ?", (user_id,))
        conn.commit()
        conn.close()

    def list_nudges(self, user_id: str) -> list[dict[str, Any]]:
        self._doctor()
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("SELECT * FROM nudges WHERE user_id = ? AND status != 'dismissed' ORDER BY created_at DESC", (user_id,))
        rows = [dict(row) for row in cur.fetchall()]
        conn.close()
        return rows

    def add_nudge(self, user_id: str, text: str, detail: str, nudge_type: str = "deadline") -> dict[str, Any]:
        self._doctor()
        n_id = str(uuid.uuid4())
        now = _now_iso()
        conn = self._connect()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO nudges (id, user_id, text, detail, type, status, created_at) VALUES (?, ?, ?, ?, ?, 'active', ?)",
            (n_id, user_id, text, detail, nudge_type, now)
        )
        conn.commit()
        conn.close()
        return {"id": n_id, "user_id": user_id, "text": text, "detail": detail, "type": nudge_type, "created_at": now, "status": "active"}

    def delete_nudges_by_type(self, user_id: str, nudge_type: str):
        self._doctor()
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("DELETE FROM nudges WHERE user_id = ? AND type = ?", (user_id, nudge_type))
        conn.commit()
        conn.close()


class SupabaseStore:
    def __init__(self, base_url: str, service_role_key: str):
        self.base_url = base_url
        self.service_role_key = service_role_key

    def _headers(self, extra: Optional[dict[str, str]] = None) -> dict[str, str]:
        headers = {
            "apikey": self.service_role_key,
            "Authorization": f"Bearer {self.service_role_key}",
            "Content-Type": "application/json",
        }
        if extra:
            headers.update(extra)
        return headers

    def _request(self, method: str, path: str, *, params: Optional[dict[str, Any]] = None, json_body: Any = None, headers: Optional[dict[str, str]] = None) -> Any:
        url = f"{self.base_url}{path}"
        with httpx.Client(timeout=20.0) as client:
            response = client.request(method, url, params=params, json=json_body, headers=self._headers(headers))
        if response.status_code >= 400:
            raise StorageError(f"Supabase request failed ({response.status_code}): {response.text[:300]}")
        if not response.text.strip():
            return None
        return response.json()

    def ensure_profile(self, user_id: str, email: str = "", full_name: str = "") -> dict[str, Any]:
        rows = self._request(
            "GET",
            "/rest/v1/profiles",
            params={"id": f"eq.{user_id}", "select": "*"},
        ) or []
        if rows:
            return rows[0]

        now = _now_iso()
        inserted = self._request(
            "POST",
            "/rest/v1/profiles",
            params={"on_conflict": "id"},
            json_body={
                "id": user_id,
                "email": email,
                "full_name": full_name or "Friend",
                "language": "English",
                "voice_gender": "female",
                "onboarding_completed": False,
                "created_at": now,
                "updated_at": now,
            },
            headers={"Prefer": "resolution=merge-duplicates,return=representation"},
        ) or []
        return inserted[0] if inserted else {
            "id": user_id,
            "email": email,
            "full_name": full_name or "Friend",
            "language": "English",
            "voice_gender": "female",
            "onboarding_completed": False,
            "created_at": now,
            "updated_at": now,
        }

    def update_profile(self, user_id: str, data: dict[str, Any]) -> dict[str, Any]:
        payload = {key: value for key, value in data.items() if key in {"email", "full_name", "language", "voice_gender", "onboarding_completed"}}
        if not payload:
            return self.ensure_profile(user_id)
        payload["updated_at"] = _now_iso()
        self._request(
            "PATCH",
            "/rest/v1/profiles",
            params={"id": f"eq.{user_id}"},
            json_body=payload,
            headers={"Prefer": "return=representation"},
        )
        return self.ensure_profile(user_id)

    def list_sessions(self, user_id: str) -> list[dict[str, Any]]:
        rows = self._request(
            "GET",
            "/rest/v1/chat_sessions",
            params={"user_id": f"eq.{user_id}", "select": "*", "order": "created_at.desc"},
        ) or []
        # Normalize: frontend expects "id" to be the session identifier
        for r in rows:
            r["id"] = r.get("session_id") or r.get("id")
        return rows

    def create_session(self, user_id: str, title: str = "New chat") -> dict[str, Any]:
        now = _now_iso()
        _session_id = str(random.randint(10000000, 99999999))
        session = self._request(
            "POST",
            "/rest/v1/chat_sessions",
            json_body={
                "session_id": _session_id,
                "user_id": user_id,
                "title": title,
                "created_at": now,
                "updated_at": now,
            },
            headers={"Prefer": "return=representation"},
        ) or []
        result = session[0]
        result["id"] = result.get("session_id") or result.get("id")
        return result

    def get_messages(self, user_id: str, session_id: str) -> list[dict[str, Any]]:
        return self._request(
            "GET",
            "/rest/v1/chat_messages",
            params={
                "user_id": f"eq.{user_id}",
                "session_id": f"eq.{session_id}",
                "select": "id,role,text,created_at",
                "order": "created_at.asc",
            },
        ) or []

    def add_message(self, user_id: str, session_id: str, role: str, text: str) -> dict[str, Any]:
        now = _now_iso()
        _msg_id = random.randint(10000000, 99999999)
        inserted = self._request(
            "POST",
            "/rest/v1/chat_messages",
            json_body={
                "id": _msg_id,
                "session_id": session_id,
                "user_id": user_id,
                "role": role,
                "text": text,
                "created_at": now,
            },
            headers={"Prefer": "return=representation"},
        ) or []
        self._request(
            "PATCH",
            "/rest/v1/chat_sessions",
            params={"session_id": f"eq.{session_id}", "user_id": f"eq.{user_id}"},
            json_body={"updated_at": now},
            headers={"Prefer": "return=minimal"},
        )
        if role == "user":
            user_messages = self._request(
                "GET",
                "/rest/v1/chat_messages",
                params={
                    "user_id": f"eq.{user_id}",
                    "session_id": f"eq.{session_id}",
                    "role": "eq.user",
                    "select": "id",
                },
            ) or []
            if len(user_messages) == 1:
                title = (text[:42].strip() + "...") if len(text.strip()) > 45 else text.strip() or "New chat"
                self._request(
                    "PATCH",
                    "/rest/v1/chat_sessions",
                    params={"session_id": f"eq.{session_id}", "user_id": f"eq.{user_id}"},
                    json_body={"title": title, "updated_at": now},
                    headers={"Prefer": "return=minimal"},
                )
        return inserted[0]

    def delete_session(self, user_id: str, session_id: str):
        self._request(
            "DELETE",
            "/rest/v1/chat_messages",
            params={"user_id": f"eq.{user_id}", "session_id": f"eq.{session_id}"},
            headers={"Prefer": "return=minimal"},
        )
        self._request(
            "DELETE",
            "/rest/v1/chat_sessions",
            params={"user_id": f"eq.{user_id}", "session_id": f"eq.{session_id}"},
            headers={"Prefer": "return=minimal"},
        )

    def get_google_integration(self, user_id: str) -> Optional[dict[str, Any]]:
        rows = self._request(
            "GET",
            "/rest/v1/google_integrations",
            params={"user_id": f"eq.{user_id}", "select": "*"},
        ) or []
        return rows[0] if rows else None

    def upsert_google_integration(self, user_id: str, data: dict[str, Any]) -> dict[str, Any]:
        now = _now_iso()
        inserted = self._request(
            "POST",
            "/rest/v1/google_integrations",
            params={"on_conflict": "user_id"},
            json_body={
                "user_id": user_id,
                "google_email": data.get("google_email"),
                "access_token": data.get("access_token"),
                "refresh_token": data.get("refresh_token"),
                "token_expiry": data.get("token_expiry"),
                "scopes": data.get("scopes") or [],
                "connected_at": data.get("connected_at") or now,
                "updated_at": now,
            },
            headers={"Prefer": "resolution=merge-duplicates,return=representation"},
        ) or []
        return inserted[0] if inserted else {}

    def delete_google_integration(self, user_id: str):
        self._request(
            "DELETE",
            "/rest/v1/google_integrations",
            params={"user_id": f"eq.{user_id}"},
            headers={"Prefer": "return=minimal"},
        )

    def list_nudges(self, user_id: str) -> list[dict[str, Any]]:
        return self._request(
            "GET",
            "/rest/v1/nudges",
            params={"user_id": f"eq.{user_id}", "status": "neq.dismissed", "select": "*", "order": "created_at.desc"},
        ) or []

    def add_nudge(self, user_id: str, text: str, detail: str, nudge_type: str = "deadline") -> dict[str, Any]:
        now = _now_iso()
        inserted = self._request(
            "POST",
            "/rest/v1/nudges",
            json_body={
                "id": str(uuid.uuid4()),
                "user_id": user_id,
                "text": text,
                "detail": detail,
                "type": nudge_type,
                "status": "active",
                "created_at": now,
            },
            headers={"Prefer": "return=representation"},
        ) or []
        return inserted[0] if inserted else {}


    def delete_nudges_by_type(self, user_id: str, nudge_type: str):
        self._request(
            "DELETE",
            f"/rest/v1/nudges?user_id=eq.{user_id}&type=eq.{nudge_type}"
        )


class ExcelStore:
    def __init__(self, excel_path: Path):
        self.path = excel_path
        self._init_excel()

    def _init_excel(self):
        try:
            import openpyxl
        except ImportError:
            return
        if not self.path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Profiles"
            ws.append(["id", "email", "full_name", "language", "voice_gender", "onboarding_completed", "created_at", "updated_at"])
            wb.save(self.path)
            
        wb = openpyxl.load_workbook(self.path)
        required = {
            "Profiles": ["id", "email", "full_name", "language", "voice_gender", "onboarding_completed", "created_at", "updated_at"],
            "Sessions": ["id", "user_id", "title", "created_at", "updated_at"],
            "Messages": ["id", "session_id", "user_id", "role", "text", "created_at"],
            "Integrations": ["user_id", "google_email", "access_token", "refresh_token", "token_expiry", "scopes_json", "connected_at", "updated_at"],
            "Nudges": ["id", "user_id", "text", "detail", "type", "created_at", "status"],
            "Memory": ["id", "user_id", "text", "created_at"],
            "Deadlines": ["id", "user_id", "title", "due_date", "source_email_id", "status", "created_at"]
        }
        updated = False
        for sheet, headers in required.items():
            if sheet not in wb.sheetnames:
                ws = wb.create_sheet(sheet)
                ws.append(headers)
                updated = True
        if updated:
            wb.save(self.path)

    def _get_sheet_data(self, sheet_name: str) -> list[dict[str, Any]]:
        import openpyxl
        from datetime import datetime, date
        wb = openpyxl.load_workbook(self.path)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
        if not headers: return []
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                # Ensure all values are JSON serializable (especially datetimes)
                serialized_row = []
                for val in row:
                    if isinstance(val, (datetime, date)):
                        serialized_row.append(val.isoformat())
                    else:
                        serialized_row.append(val)
                rows.append(dict(zip(headers, serialized_row)))
        return rows

    def _save_sheet_data(self, sheet_name: str, data: list[dict[str, Any]]):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(self.path)
        except PermissionError:
            raise StorageError(f"DATABASE LOCKED: Please CLOSE '{self.path.name}' in Excel so Saathi can write to it.")
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            if data:
                headers = list(data[0].keys())
                ws.append(headers)
            else:
                try: wb.save(self.path)
                except PermissionError: raise StorageError(f"DATABASE LOCKED: Please CLOSE '{self.path.name}' in Excel.")
                return
        else:
            ws = wb[sheet_name]
        
        # Clear existing rows except headers
        while ws.max_row > 1:
            ws.delete_rows(2)
        
        headers = [cell.value for cell in ws[1]]
        if not headers and data:
            headers = list(data[0].keys())
            ws.append(headers)

        for item in data:
            row = [item.get(h) for h in headers]
            ws.append(row)
        
        try:
            wb.save(self.path)
        except PermissionError:
            raise StorageError(f"DATABASE LOCKED: Please CLOSE '{self.path.name}' in Excel so Saathi can save your data.")

    def ensure_profile(self, user_id: str, email: str = "", full_name: str = "") -> dict[str, Any]:
        profiles = self._get_sheet_data("Profiles")
        for p in profiles:
            if p["id"] == user_id:
                # Handle Excel's loosely typed values (0, 1, "0", "1", True, False)
                val = p.get("onboarding_completed")
                p["onboarding_completed"] = val in [1, "1", True, "True"]
                return p
        now = _now_iso()
        new_profile = {
            "id": user_id,
            "email": email,
            "full_name": full_name or "Friend",
            "language": "English",
            "voice_gender": "female",
            "onboarding_completed": 0,
            "created_at": now,
            "updated_at": now,
        }
        profiles.append(new_profile)
        self._save_sheet_data("Profiles", profiles)
        new_profile["onboarding_completed"] = False
        return new_profile

    def profile_exists(self, user_id: str) -> bool:
        # If the Excel file exists and contains the user record, they exist.
        if not self.path.exists(): return False
        profiles = self._get_sheet_data("Profiles")
        return any(p["id"] == user_id for p in profiles)

    def update_profile(self, user_id: str, data: dict[str, Any]) -> dict[str, Any]:
        profiles = self._get_sheet_data("Profiles")
        updated = False
        for p in profiles:
            if p["id"] == user_id:
                for k, v in data.items():
                    if k in p:
                        p[k] = v
                p["updated_at"] = _now_iso()
                updated = True
                break
        if updated:
            self._save_sheet_data("Profiles", profiles)
        return self.ensure_profile(user_id)

    def list_sessions(self, user_id: str) -> list[dict[str, Any]]:
        sessions = self._get_sheet_data("Sessions")
        return [s for s in sessions if s["user_id"] == user_id]

    def create_session(self, user_id: str, title: str = "New chat") -> dict[str, Any]:
        session_id = str(uuid.uuid4())
        now = _now_iso()
        session = {
            "id": session_id,
            "user_id": user_id,
            "title": title,
            "created_at": now,
            "updated_at": now,
        }
        sessions = self._get_sheet_data("Sessions")
        sessions.append(session)
        self._save_sheet_data("Sessions", sessions)
        return session

    def get_messages(self, user_id: str, session_id: str) -> list[dict[str, Any]]:
        msgs = self._get_sheet_data("Messages")
        return [m for m in msgs if m["user_id"] == user_id and m["session_id"] == session_id]

    def add_message(self, user_id: str, session_id: str, role: str, text: str) -> dict[str, Any]:
        now = _now_iso()
        msg_id = str(uuid.uuid4())
        msg = {
            "id": msg_id,
            "session_id": session_id,
            "user_id": user_id,
            "role": role,
            "text": text,
            "created_at": now,
        }
        msgs = self._get_sheet_data("Messages")
        msgs.append(msg)
        self._save_sheet_data("Messages", msgs)
        
        # Update session title if first user message
        if role == "user":
            user_msgs = [m for m in msgs if m["user_id"] == user_id and m["session_id"] == session_id and m["role"] == "user"]
            if len(user_msgs) == 1:
                title = (text[:42].strip() + "...") if len(text.strip()) > 45 else text.strip() or "New chat"
                sessions = self._get_sheet_data("Sessions")
                for s in sessions:
                    if s["id"] == session_id and s["user_id"] == user_id:
                        s["title"] = title
                        s["updated_at"] = now
                        break
                self._save_sheet_data("Sessions", sessions)
        return msg

    def delete_session(self, user_id: str, session_id: str):
        msgs = [m for m in self._get_sheet_data("Messages") if not (m["user_id"] == user_id and m["session_id"] == session_id)]
        self._save_sheet_data("Messages", msgs)
        sessions = [s for s in self._get_sheet_data("Sessions") if not (s["user_id"] == user_id and s["id"] == session_id)]
        self._save_sheet_data("Sessions", sessions)

    def get_google_integration(self, user_id: str) -> Optional[dict[str, Any]]:
        ints = self._get_sheet_data("Integrations")
        for i in ints:
            if i["user_id"] == user_id:
                i["scopes"] = json.loads(i.get("scopes_json") or "[]")
                return i
        return None

    def upsert_google_integration(self, user_id: str, data: dict[str, Any]) -> dict[str, Any]:
        ints = self._get_sheet_data("Integrations")
        found = False
        now = _now_iso()
        for i in ints:
            if i["user_id"] == user_id:
                i.update({
                    "google_email": data.get("google_email"),
                    "access_token": data.get("access_token"),
                    "refresh_token": data.get("refresh_token"),
                    "token_expiry": data.get("token_expiry"),
                    "scopes_json": json.dumps(data.get("scopes") or []),
                    "updated_at": now,
                })
                found = True
                break
        if not found:
            ints.append({
                "user_id": user_id,
                "google_email": data.get("google_email"),
                "access_token": data.get("access_token"),
                "refresh_token": data.get("refresh_token"),
                "token_expiry": data.get("token_expiry"),
                "scopes_json": json.dumps(data.get("scopes") or []),
                "connected_at": now,
                "updated_at": now,
            })
        self._save_sheet_data("Integrations", ints)
        return self.get_google_integration(user_id) or {}

    def delete_google_integration(self, user_id: str):
        ints = [i for i in self._get_sheet_data("Integrations") if i["user_id"] != user_id]
        self._save_sheet_data("Integrations", ints)

    def list_nudges(self, user_id: str) -> list[dict[str, Any]]:
        nudges = self._get_sheet_data("Nudges")
        return [n for n in nudges if n["user_id"] == user_id and n.get("status") != "dismissed"]

    def add_nudge(self, user_id: str, text: str, detail: str, nudge_type: str = "deadline") -> dict[str, Any]:
        now = _now_iso()
        nudge_id = str(uuid.uuid4())
        nudge = {
            "id": nudge_id,
            "user_id": user_id,
            "text": text,
            "detail": detail,
            "type": nudge_type,
            "created_at": now,
            "status": "active"
        }
        nudges = self._get_sheet_data("Nudges")
        nudges.append(nudge)
        self._save_sheet_data("Nudges", nudges)
        return nudge


    def delete_nudges_by_type(self, user_id: str, nudge_type: str):
        nudges = [n for n in self._get_sheet_data("Nudges") if not (str(n["user_id"]) == str(user_id) and n["type"] == nudge_type)]
        self._save_sheet_data("Nudges", nudges)

    # --- Cognitive Methods (New v29+) ---
    def list_memory(self, user_id: str) -> list[dict[str, Any]]:
        mems = self._get_sheet_data("Memory")
        return [m for m in mems if str(m["user_id"]) == str(user_id)]

    def add_memory(self, user_id: str, text: str) -> dict[str, Any]:
        now = _now_iso()
        m_id = str(uuid.uuid4())
        mem = {"id": m_id, "user_id": user_id, "text": text, "created_at": now}
        mems = self._get_sheet_data("Memory")
        mems.append(mem)
        self._save_sheet_data("Memory", mems)
        return mem

    def delete_memory(self, user_id: str, memory_id: str):
        mems = [m for m in self._get_sheet_data("Memory") if not (str(m["user_id"]) == str(user_id) and str(m["id"]) == str(memory_id))]
        self._save_sheet_data("Memory", mems)

    def list_deadlines(self, user_id: str) -> list[dict[str, Any]]:
        deadlines = self._get_sheet_data("Deadlines")
        return [d for d in deadlines if str(d["user_id"]) == str(user_id)]

    def add_deadline(self, user_id: str, title: str, due_date: str, source_email_id: str = "") -> dict[str, Any]:
        now = _now_iso()
        d_id = str(uuid.uuid4())
        deadline = {
            "id": d_id, "user_id": user_id, "title": title, 
            "due_date": due_date, "source_email_id": source_email_id, 
            "status": "pending", "created_at": now
        }
        deadlines = self._get_sheet_data("Deadlines")
        # De-duplicate
        if any(d["title"] == title and d["user_id"] == user_id for d in deadlines): return deadline
        deadlines.append(deadline)
        self._save_sheet_data("Deadlines", deadlines)
        return deadline

class StorageService:
    def __init__(self):
        self.local = LocalStore(LOCAL_DB_FILE)
        self.remote = SupabaseStore(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY) if supabase_enabled() else None
        self.mode = "excel"
        
        # User-Specific Excel Registry
        self.excel_stores = {} # Cache of user-specific ExcelStore instances
        self.storage_dir = Path("user_vaults")
        self.storage_dir.mkdir(exist_ok=True)

    def _get_store(self, user_id: str):
        # We prioritize Multi-Tenant Excel as requested
        if user_id not in self.excel_stores:
            safe_id = "".join([c if c.isalnum() else "_" for c in str(user_id)])
            excel_path = self.storage_dir / f"saathi_storage_{safe_id}.xlsx"
            self.excel_stores[user_id] = ExcelStore(excel_path)
        return self.excel_stores[user_id]

    def _backend(self, user_id: str):
        # Priotize the user's personal Excel vault
        return self._get_store(user_id)

    # --- Proxy Layer (Redirecting to User-Specific Vaults) ---
    def ensure_profile(self, user_id: str, email: str = "", full_name: str = "") -> dict[str, Any]:
        return self._backend(user_id).ensure_profile(user_id, email, full_name)

    def profile_exists(self, user_id: str) -> bool:
        return self._backend(user_id).profile_exists(user_id)

    def update_profile(self, user_id: str, data: dict[str, Any]) -> dict[str, Any]:
        return self._backend(user_id).update_profile(user_id, data)

    def list_sessions(self, user_id: str) -> list[dict[str, Any]]:
        return self._backend(user_id).list_sessions(user_id)

    def create_session(self, user_id: str, title: str = "New chat") -> dict[str, Any]:
        return self._backend(user_id).create_session(user_id, title)

    def get_messages(self, user_id: str, session_id: str) -> list[dict[str, Any]]:
        return self._backend(user_id).get_messages(user_id, session_id)

    def add_message(self, user_id: str, session_id: str, role: str, text: str) -> dict[str, Any]:
        return self._backend(user_id).add_message(user_id, session_id, role, text)

    def delete_session(self, user_id: str, session_id: str):
        return self._backend(user_id).delete_session(user_id, session_id)

    def list_memory(self, user_id: str) -> list[dict[str, Any]]:
        return self._backend(user_id).list_memory(user_id)

    def add_memory(self, user_id: str, text: str) -> dict[str, Any]:
        return self._backend(user_id).add_memory(user_id, text)

    def delete_memory(self, user_id: str, memory_id: str):
        return self._backend(user_id).delete_memory(user_id, memory_id)

    def list_deadlines(self, user_id: str) -> list[dict[str, Any]]:
        return self._backend(user_id).list_deadlines(user_id)

    def add_deadline(self, user_id: str, title: str, due_date: str, source_email_id: str = "") -> dict[str, Any]:
        return self._backend(user_id).add_deadline(user_id, title, due_date, source_email_id)

    def delete_deadline(self, user_id: str, deadline_id: str):
        return self._backend(user_id).delete_deadline(user_id, deadline_id)

    def get_google_integration(self, user_id: str) -> Optional[dict[str, Any]]:
        return self._backend(user_id).get_google_integration(user_id)

    def upsert_google_integration(self, user_id: str, data: dict[str, Any]) -> dict[str, Any]:
        return self._backend(user_id).upsert_google_integration(user_id, data)

    def delete_google_integration(self, user_id: str):
        return self._backend(user_id).delete_google_integration(user_id)

    def list_nudges(self, user_id: str) -> list[dict[str, Any]]:
        return self._backend(user_id).list_nudges(user_id)

    def add_nudge(self, user_id: str, text: str, detail: str, nudge_type: str = "deadline") -> dict[str, Any]:
        return self._backend(user_id).add_nudge(user_id, text, detail, nudge_type)

    def delete_nudges_by_type(self, user_id: str, nudge_type: str):
        return self._backend(user_id).delete_nudges_by_type(user_id, nudge_type)


storage = StorageService()


async def verify_access_token(access_token: str) -> Optional[dict[str, Any]]:
    if not access_token:
        return None
    if not (SUPABASE_URL and SUPABASE_ANON_KEY):
        return {
            "id": "demo-user",
            "email": "demo@saathi.local",
            "user_metadata": {"full_name": "Demo User"},
            "app_metadata": {"provider": "demo"},
        }
    headers = {
        "apikey": SUPABASE_ANON_KEY,
        "Authorization": f"Bearer {access_token}",
    }
    async with httpx.AsyncClient(timeout=15.0) as client:
        response = await client.get(f"{SUPABASE_URL}/auth/v1/user", headers=headers)
    if response.status_code >= 400:
        print(f"SUPABASE AUTH ERROR: {response.text}")
        return None
    return response.json()
